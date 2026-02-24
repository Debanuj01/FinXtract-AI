import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
MAROON    = "741B47"
WHITE     = "FFFFFF"
GREY_BG   = "F2F2F2"
GREEN_BG  = "E2EFDA"
BLUE_BG   = "DDEBF7"
YELLOW_BG = "FFF2CC"
BLACK     = "000000"
GREY_TXT  = "888888"

_thin   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _font(bold=False, color=BLACK, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


# ── Row classification ─────────────────────────────────────────────────────────
TOTAL_KW = [
    "total", "gross profit", "ebitda", "profit before", "profit after",
    "profit for", "earnings after", "net profit", "tax expense",
    "total expenses", "total income", "total revenue", "total employee",
    "total finance", "total other", "total purchases", "total comprehensive"
]

SECTION_KW = [
    "expenses", "revenue from operations", "other comprehensive income",
    "earnings per share", "paid-up equity"
]

EXCEPTIONAL_KW = ["exceptional", "discontinued"]


def classify(name):
    nl = name.lower().strip()
    if any(nl == s for s in SECTION_KW):
        return "section"
    if any(kw in nl for kw in TOTAL_KW):
        return "total"
    if any(kw in nl for kw in EXCEPTIONAL_KW):
        return "exceptional"
    return "normal"


# ── Main generator ─────────────────────────────────────────────────────────────
def generate_excel(data, metadata, output_folder=None):

    if output_folder is None:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        output_folder = os.path.join(base, "output")

    os.makedirs(output_folder, exist_ok=True)

    # ── Extract data fields (declared ONCE) ───────────────────────────────────
    years        = data.get("years", [])
    line_items   = data.get("line_items", [])
    currency     = data.get("currency", "INR")
    units        = data.get("units", "crores")
    company_name = data.get("company_name", "Financial Statement").upper()
    num_years    = len(years)
    last_col     = get_column_letter(2 + num_years)

    # ── Build output filename from company name ────────────────────────────────
    safe_name   = "".join(c for c in company_name if c.isalnum() or c in (" ", "-", "_", ".")).strip()
    file_name   = f"{safe_name}.xlsx" if safe_name else "Financial_Statement.xlsx"
    output_path = os.path.join(output_folder, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    # ── Row 1: Title banner ────────────────────────────────────────────────────
    ws.merge_cells(f"B1:{last_col}1")
    c = ws["B1"]
    c.value     = f"{company_name}  |  Consolidated Income Statement  |  ₹ in {units}"
    c.font      = _font(bold=True, color=WHITE, size=12)
    c.fill      = _fill(MAROON)
    c.alignment = _align("center")
    c.border    = _BORDER
    ws.row_dimensions[1].height = 24

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    headers = ["Particulars"] + [str(y) for y in years]
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(row=2, column=ci, value=h)      # ← row=2 (not row=3)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.fill      = _fill(MAROON)
        c.alignment = _align("center")
        c.border    = _BORDER
    ws.row_dimensions[2].height = 20

    # ── Data rows start at row 3 ───────────────────────────────────────────────
    row_map = {}
    cur_row = 3                                      # ← starts at 3 (not 4)

    for item in line_items:
        name   = item.get("name", "").strip()
        values = item.get("values", [])
        if not name:
            continue

        nl   = name.lower()
        kind = classify(name)

        if kind == "section":
            bg, bold, indent = None, True, 0
        elif kind == "total":
            bg, bold, indent = GREY_BG, True, 0
        elif kind == "exceptional":
            bg, bold, indent = YELLOW_BG, False, 2
        else:
            bg, bold, indent = None, False, 2

        if "ebitda" in nl:
            bg, bold = GREEN_BG, True
        if "profit for the year" in nl or "profit after tax" in nl:
            bg, bold = BLUE_BG, True
        if "profit before tax" in nl:
            bg, bold = BLUE_BG, True

        if bg is None and cur_row % 2 == 0:
            bg = "FAFAFA"

        # Label cell
        lc = ws.cell(row=cur_row, column=2, value=name)
        lc.font      = _font(bold=bold, size=10)
        lc.alignment = _align("left", wrap=True, indent=indent)
        lc.border    = _BORDER
        if bg:
            lc.fill = _fill(bg)

        # Value cells
        for ci_off in range(num_years):
            col_idx = 3 + ci_off
            val     = values[ci_off] if ci_off < len(values) else None
            vc      = ws.cell(row=cur_row, column=col_idx)

            if val is None:
                vc.value = "–"
                vc.font  = _font(color=GREY_TXT, size=10)
            elif isinstance(val, (int, float)):
                vc.value         = val
                vc.number_format = '#,##0.00'
                vc.font          = _font(bold=bold, size=10)
            else:
                vc.value = val
                vc.font  = _font(bold=bold, size=10)

            vc.alignment = _align("center")
            vc.border    = _BORDER
            if bg:
                vc.fill = _fill(bg)

        row_map[nl] = cur_row
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    # ── Key Metrics section ────────────────────────────────────────────────────
    cur_row += 1
    ws.merge_cells(f"B{cur_row}:{last_col}{cur_row}")
    hc = ws.cell(row=cur_row, column=2, value="KEY METRICS  (Live Excel Formulas)")
    hc.font      = _font(bold=True, color=WHITE, size=10)
    hc.fill      = _fill(MAROON)
    hc.alignment = _align("center")
    hc.border    = _BORDER
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    def find_row(keyword):
        for k, r in row_map.items():
            if keyword in k:
                return r
        return None

    rev_r = find_row("revenue from operations")
    oth_r = find_row("other income")
    dep_r = find_row("depreciation")
    fin_r = find_row("finance costs")
    pbt_r = find_row("profit before tax")
    pat_r = find_row("profit for the year from continuing")
    exp_r = find_row("total expenses")

    metrics = []
    if rev_r and oth_r:
        metrics.append(("Total Revenue (Ops + Other Income)",
                         lambda c, r=rev_r, o=oth_r: f"={c}{r}+{c}{o}",
                         GREY_BG, True, False))
    if pbt_r and fin_r and dep_r:
        metrics.append(("EBITDA  (PBT + Finance Costs + D&A)",
                         lambda c, p=pbt_r, f=fin_r, d=dep_r: f"={c}{p}+{c}{f}+{c}{d}",
                         GREEN_BG, True, False))
    if pat_r and rev_r:
        metrics.append(("PAT Margin %",
                         lambda c, p=pat_r, r=rev_r: f"={c}{p}/{c}{r}",
                         BLUE_BG, False, True))
    if pbt_r and rev_r:
        metrics.append(("PBT Margin %",
                         lambda c, p=pbt_r, r=rev_r: f"={c}{p}/{c}{r}",
                         BLUE_BG, False, True))
    if exp_r and rev_r:
        metrics.append(("Cost Ratio (Total Expenses / Revenue)",
                         lambda c, e=exp_r, r=rev_r: f"={c}{e}/{c}{r}",
                         GREY_BG, False, True))

    for (mname, formula_fn, bg, bold, is_pct) in metrics:
        lc = ws.cell(row=cur_row, column=2, value=mname)
        lc.font      = _font(bold=bold, size=10)
        lc.fill      = _fill(bg)
        lc.alignment = _align("left")
        lc.border    = _BORDER

        for ci_off in range(num_years):
            col_ltr = get_column_letter(3 + ci_off)
            vc = ws.cell(row=cur_row, column=3+ci_off, value=formula_fn(col_ltr))
            vc.font          = _font(bold=bold, size=10)
            vc.fill          = _fill(bg)
            vc.alignment     = _align("center")
            vc.border        = _BORDER
            vc.number_format = '0.00%' if is_pct else '#,##0.00'

        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 42
    for ci_off in range(num_years):
        ws.column_dimensions[get_column_letter(3+ci_off)].width = 16

    ws.freeze_panes = "C3"                           # ← freeze at C3 (headers on row 2)

    # ── Sheet 2: Metadata ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Metadata")
    for ci, h in enumerate(["Field", "Value"], start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.fill      = _fill(MAROON)
        c.alignment = _align("center")
        c.border    = _BORDER

    EXCLUDE_KEYS = {"confidence", "method", "accuracy"}
    extra_meta   = [(k, str(v)) for k, v in metadata.items() if k.lower() not in EXCLUDE_KEYS]

    meta_rows = [
        ("Company",   company_name.title()),
        ("Currency",  currency),
        ("Units",     units),
        ("Periods",   ", ".join(str(y) for y in years)),
        ("Statement", "Consolidated Audited Financial Results"),
    ] + extra_meta

    for ri, (field, value) in enumerate(meta_rows, start=2):
        ws2.cell(row=ri, column=1, value=field).font   = _font(bold=True, size=10)
        ws2.cell(row=ri, column=1).border              = _BORDER
        ws2.cell(row=ri, column=1).alignment           = _align("left")
        ws2.cell(row=ri, column=2, value=str(value)).font = _font(size=10)
        ws2.cell(row=ri, column=2).border              = _BORDER
        ws2.cell(row=ri, column=2).alignment           = _align("left")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55

    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    return output_path